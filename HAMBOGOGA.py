# tkinter
from tkinter import *
from tkinter import font
from tkinter import ttk

# webbrowser
import webbrowser

# excel
from openpyxl import load_workbook

# xml
import urllib
from urllib.request import urlopen
from urllib.parse import urlencode, unquote, quote_plus
import xml.etree.ElementTree as ET

# re
import re

# date
import datetime

# cef
import threading
import sys
from cefpython3 import cefpython as cef

# time module
import cpptime

# telegram
import time
import sqlite3
import telepot
from pprint import pprint
from bs4 import BeautifulSoup
import traceback

import noti

class HAMBOGOGA :
    def __init__(self) :
        self.window_main = Tk()
        self.window_main.geometry("1200x900")

        # data
        self.ReadExelFile()

        # Init
        self.InitAppTitle()
        self.InitLocationListBox()
        self.InitSearchLayout()
        self.InitInfoButton()
        self.InitInfoCanvas()
        self.InitAPIInfo()
        self.InitMap()
        self.InitBot()

        self.window_main.mainloop()

    # read file
    def ReadExelFile(self) :
        self.LocalData_wb = load_workbook("LocalData.xlsx", data_only= True)
        self.LocalData_ws = self.LocalData_wb['Sheet1']

        # A : 시도
        # B : 시군구
        # C : 읍면동
        # D : x
        # E : y
        # F : 경도
        # G : 위도

    # init
    def InitAppTitle(self) :
        self.photo_logo = PhotoImage(file= "title_logo.png")
        self.button_AppTitle = Button(self.window_main, text= "HAMBOGOGA", image= self.photo_logo, relief= FLAT, command= self.Clicked_Title)    # height 1 : 25?
        self.button_AppTitle.place(x= 13, y= 15)

    def InitLocationListBox(self) :
        self.frame_Location = Frame(self.window_main, width= 33, height= 10)
        self.frame_Location.place(x= 15, y= 80)

        self.listbox_Location = Listbox(self.frame_Location, width= 33, height= 10)
        self.listbox_Location.pack(side= 'left', fill= 'y')

        self.scrollbar_Location = Scrollbar(self.frame_Location, command= self.listbox_Location.yview)
        self.scrollbar_Location.pack(side= 'right', fill= 'y')

        self.listbox_Location.config(yscrollcommand= self.scrollbar_Location.set)
        self.listbox_Location.event_generate("<<ListboxSelect>>")
        self.listbox_Location.bind("<<ListboxSelect>>", self.Selected_ListBox)
        
        count = 1
        for row in self.LocalData_ws.rows :
            self.listbox_Location.insert(count, row[7].value)
            count += 1

    def InitSearchLayout(self) :
        self.entry_Search = Entry(self.window_main, width= 25)
        self.entry_Search.insert(0, "정왕1동")
        self.entry_Search.place(x= 315, y= 84)

        self.button_Search = Button(self.window_main, text= "검색", width=5, command= self.Clicked_Search)
        self.button_Search.place(x= 535, y= 80)

    def InitInfoButton(self) :
        # pm
        self.button_PM = Button(self.window_main, text= "미세먼지 정보", width= 33, height= 1, command= self.Show_PM)
        self.button_PM.place(x= 311, y= 130)
        
        # weather
        self.butto_Weather = Button(self.window_main, text= "날씨 정보", width= 33, height= 1, command= self.Show_Weather)
        self.butto_Weather.place(x= 311, y= 170)
        
        # stock
        stocks = ('넥슨게임즈', '펄어비스', '크래프톤', '넷마블', '데브시스터즈')

        self.selected_stock = StringVar()
        self.combobox_stock = ttk.Combobox(self.window_main, values= stocks, width= 31, height= 1, textvariable= self.selected_stock)
        self.combobox_stock.set("Select a stock")
        self.combobox_stock.place(x= 313, y= 215)

        self.butto_Stock = Button(self.window_main, text= "주식 정보", width= 33, height= 1, command= self.Search_Stock)
        self.butto_Stock.place(x= 311, y= 250)

    def InitInfoCanvas(self) :
        # 1
        self.canvas_Info = Canvas(self.window_main, width= 567, height= 300, bg= 'white')
        self.canvas_Info.place(x= 15, y= 300)

        # 2
        self.canvas_Info2 = Canvas(self.window_main, width= 567, height= 260, bg= 'white')
        self.canvas_Info2.place(x= 15, y= 620)

    def InitAPIInfo(self) :
        self.serviceKey = "QHpOtm0e0OwX2cl8WWXuWGQoaOkbXfXYjF60tquzusBWCg3488dfLbTLACxkPJr1EyPxSYd27VCOUh6ZS+RhPQ=="

    def InitMap(self) :
        self.frame_Map = Frame(self.window_main, width= 570, height= 870)
        self.frame_Map.place(x= 615, y= 15)
        thread = threading.Thread(target= self.Show_Map, args= (self.frame_Map,))
        thread.daemon = True
        thread.start()

    def InitBot(self) :
        self.bot = telepot.Bot(noti.TOKEN)
        pprint( self.bot.getMe() )

        self.bot.message_loop(self.handle)

        print('Listening...')

    # about map
    def Show_Map(self, frame) :
        sys.excepthook = cef.ExceptHook
        window_info = cef.WindowInfo(frame.winfo_id())
        window_info.SetAsChild(frame.winfo_id(), [0, 0, 570, 870])
        cef.Initialize()
        browser = cef.CreateBrowserSync(window_info, url= 'https://www.misemise.co.kr/')
        cef.MessageLoop()

    # about title
    def Clicked_Title(self) :
        webbrowser.open("https://www.tukorea.ac.kr")

    # about listbox
    def Selected_ListBox(self, event) :
        try :
            self.entry_Search.insert(0, self.LocalData_ws.cell(event.widget.curselection()[0] + 1, 3).value)
            self.entry_Search.delete(0, 'end')
            self.entry_Search.insert(0, self.LocalData_ws.cell(event.widget.curselection()[0] + 1, 3).value)
        except :
            pass

    # about search
    def Clicked_Search(self) :
        # PM
        callbackURL = "http://apis.data.go.kr/B552584/ArpltnInforInqireSvc/getMsrstnAcctoRltmMesureDnsty"
        params = '?' + urlencode({
            quote_plus("serviceKey"): self.serviceKey,
            quote_plus("returnType"): "xml",
            quote_plus("numOfRows"): "10",
            quote_plus("pageNo"): "1",
            quote_plus("stationName"): re.sub('\d+', '', self.entry_Search.get()),
            quote_plus("dataTerm"): "DAILY",
            quote_plus("ver"): "1.4"
        })

        url = callbackURL + params
        response_body = urlopen(url).read()
        root = ET.fromstring(response_body.decode('utf-8'))
        items = root.findall(".//item")

        if not items :
            location = ""

            for row in self.LocalData_ws.rows :
                if row[2].value == self.entry_Search.get() :
                    location = row[1].value
                    break

            callbackURL = "http://apis.data.go.kr/B552584/ArpltnInforInqireSvc/getMsrstnAcctoRltmMesureDnsty"
            params = '?' + urlencode({
                quote_plus("serviceKey"): self.serviceKey,
                quote_plus("returnType"): "xml",
                quote_plus("numOfRows"): "10",
                quote_plus("pageNo"): "1",
                quote_plus("stationName"): location,
                quote_plus("dataTerm"): "DAILY",
                quote_plus("ver"): "1.4"
            })

            url = callbackURL + params
            response_body = urlopen(url).read()
            root = ET.fromstring(response_body.decode('utf-8'))
            items = root.findall(".//item")
            

        self.allinfo_PM = []
        for item in items :
            info_PM = {
                "dataTime": item.findtext("dataTime"),
                "stationName": item.findtext("stationName"),
                "pm10": item.findtext("pm10Value"),
                "pm10Grade": item.findtext("pm10Grade"),
                "pm25": item.findtext("pm25Value"),
                "pm25Grade": item.findtext("pm25Grade"),
                "o3": item.findtext("o3Value"),
                "o3Grade": item.findtext("o3Grade"),
                "no2": item.findtext("no2Value"),
                "no2Grade": item.findtext("no2Grade"),
                "co": item.findtext("coValue"),
                "coGrade": item.findtext("coGrade"),
                "so2": item.findtext("so2Value"),
                "so2Grade": item.findtext("so2Grade")
            }

            self.allinfo_PM.append(info_PM)

        self.Show_PM()

        # Weather fcst
        # today = re.sub(r'[^0-9]', '', str(datetime.date.today()))   # date
        today = cpptime.date_bn(0)

        # x, y
        nx, ny = -1, -1

        for row in self.LocalData_ws.rows :
            if row[2].value == self.entry_Search.get() :
                nx, ny = row[3].value, row[4].value
                break

        if nx != -1 and ny != -1 :
            callbackURL = "http://apis.data.go.kr/1360000/VilageFcstInfoService_2.0/getVilageFcst"
            params = '?' + urlencode({
                quote_plus("serviceKey"): self.serviceKey,
                quote_plus("pageNo"): "1",
                quote_plus("numOfRows"): "500",
                quote_plus("dataType"): "XML",
                quote_plus("base_date"): today,
                quote_plus("base_time"): "0200",
                quote_plus("nx"): nx,
                quote_plus("ny"): ny
            })

            url = callbackURL + params
            response_body = urlopen(url).read()
            root = ET.fromstring(response_body.decode('utf-8'))
            items = root.findall(".//item")

            self.allinfo_Weather = dict()
            for item in items :
                if item.findtext("fcstDate") == item.findtext("baseDate") :
                    fcsttime = item.findtext("fcstTime")

                    if fcsttime not in self.allinfo_Weather :
                        self.allinfo_Weather[fcsttime] = dict()
                    self.allinfo_Weather[fcsttime][item.findtext("category")] = item.findtext("fcstValue")

        # Weather now
        nowtime = datetime.datetime.now().strftime("%H")

        if nowtime[0] == '0' :
            nowtime = nowtime[0] + str(eval(nowtime[1]) - 1) + "00"
        else :
            nowtime = str(eval(nowtime) - 1) + "00"

        # x, y
        nx, ny = -1, -1

        for row in self.LocalData_ws.rows :
            if row[2].value == self.entry_Search.get() :
                nx, ny = row[3].value, row[4].value
                break

        if nx != -1 and ny != -1 :
            callbackURL = "http://apis.data.go.kr/1360000/VilageFcstInfoService_2.0/getUltraSrtNcst"
            params = '?' + urlencode({
                quote_plus("serviceKey"): self.serviceKey,
                quote_plus("pageNo"): "1",
                quote_plus("numOfRows"): "500",
                quote_plus("dataType"): "XML",
                quote_plus("base_date"): today,
                quote_plus("base_time"): nowtime,
                quote_plus("nx"): nx,
                quote_plus("ny"): ny
            })

            url = callbackURL + params
            response_body = urlopen(url).read()
            root = ET.fromstring(response_body.decode('utf-8'))
            items = root.findall(".//item")

            self.nowinfo_Weather = dict()

            self.nowinfo_Weather["baseDate"] = today
            self.nowinfo_Weather["baseTime"] = nowtime

            for item in items :
                self.nowinfo_Weather[item.findtext("category")] = item.findtext("obsrValue")

    # PMInfo
    def Show_PM(self) :
        self.Show_PMInfo1()
        self.Show_PMInfo2(self.allinfo_PM)

    def Show_PMInfo1(self) :
        self.canvas_Info.delete('all')

        width = 567
        height = 300

        min = eval(self.allinfo_PM[0]['pm10'])
        max = eval(self.allinfo_PM[0]['pm10'])

        for i in range(len(self.allinfo_PM)) :
            temp = eval(self.allinfo_PM[i]['pm10'])

            if temp < min :
                min = temp

            if temp > max :
                max = temp

        count = len(self.allinfo_PM)

        x_width = width // (count + 1)
        x_gap = x_width // 5
        x_start = x_width - 10
        y_gap = 50
        y_start = 25
        y_width = height - y_gap * 2 - y_start
        y_stretch = y_width / (max - min)

        for i in range(count) :
            x0 = x_start + i * x_width
            x1 = x0 + x_gap
            y1 = height - y_gap
            y0 = y1 - y_start - (eval(self.allinfo_PM[count - i - 1]['pm10']) - min) * y_stretch

            self.canvas_Info.create_text(x0 + x_gap + 3, y1 + 10, text= self.allinfo_PM[count - i - 1]['dataTime'][11:13])
            self.canvas_Info.create_rectangle(x0, y0, x1, y1, fill= 'red')
            self.canvas_Info.create_text(x0 + x_gap / 2, y0 - 10, text= self.allinfo_PM[count - i - 1]['pm10'])

        # PM25
        min = eval(self.allinfo_PM[0]['pm25'])
        max = eval(self.allinfo_PM[0]['pm25'])

        for i in range(len(self.allinfo_PM)) :
            temp = eval(self.allinfo_PM[i]['pm25'])

            if temp < min :
                min = temp

            if temp > max :
                max = temp
                
        x_start = x_start + 15
        y_stretch = y_width / (max - min)

        for i in range(count) :
            x0 = x_start + i * x_width
            x1 = x0 + x_gap
            y1 = height - y_gap
            y0 = y1 - y_start - (eval(self.allinfo_PM[count - i - 1]['pm25']) - min) * y_stretch

            self.canvas_Info.create_rectangle(x0, y0, x1, y1, fill= 'blue')
            self.canvas_Info.create_text(x0 + x_gap / 2, y0 - 10, text= self.allinfo_PM[count - i - 1]['pm25'])

    def Show_PMInfo2(self, info_pm) :
        self.canvas_Info2.delete('all')

        self.Create_Rectangle_In_Canvas("측정 시간 : " + info_pm[0]["dataTime"], 0, 0, 0)
        self.Create_Rectangle_In_Canvas("측정소 : " + info_pm[0]["stationName"], 0, 1, 0)
        self.Create_Rectangle_In_Canvas("미세먼지 : " + info_pm[0]["pm10"] + " ㎍/㎥", 1, 0, eval(info_pm[0]["pm10Grade"]))
        self.Create_Rectangle_In_Canvas("초미세먼지 : " + info_pm[0]["pm25"] + " ㎍/㎥", 1, 1, eval(info_pm[0]["pm25Grade"]))
        self.Create_Rectangle_In_Canvas("오존 : " + info_pm[0]["o3"] + " ppm", 2, 0, eval(info_pm[0]["o3Grade"]))
        self.Create_Rectangle_In_Canvas("이산화질소 : " + info_pm[0]["no2"] + " ppm", 2, 1, eval(info_pm[0]["no2Grade"]))
        self.Create_Rectangle_In_Canvas("일산화탄소 : " + info_pm[0]["co"] + " ppm", 3, 0, eval(info_pm[0]["coGrade"]))
        self.Create_Rectangle_In_Canvas("아황산가스 : " + info_pm[0]["so2"] + " ppm", 3, 1, eval(info_pm[0]["so2Grade"]))

        # grade
        self.Create_Rectangle_In_Canvas("좋음", 11, 0, 1)
        self.Create_Rectangle_In_Canvas("보통", 11, 1, 2)
        self.Create_Rectangle_In_Canvas("나쁨", 12, 0, 3)
        self.Create_Rectangle_In_Canvas("매우 나쁨", 12, 1, 4)

    def Create_Rectangle_In_Canvas(self, intext, row, col, grade) :
        color = ["white", "SkyBlue1", "pale green", "yellow2", "IndianRed1"]

        test = self.canvas_Info2.create_text(142 + 284 * col, 10 + 20 * row, text= intext)
        rect = self.canvas_Info2.create_rectangle(self.canvas_Info2.bbox(test), fill= color[grade])
        x1, y1, x2, y2 = self.canvas_Info2.coords(rect)
        self.canvas_Info2.coords(rect, 284 * col, y1, 284 * col + 284, y2)
        self.canvas_Info2.tag_lower(rect, test)

    # Weather Info
    def Show_Weather(self) :
        self.Show_WeatherInfo1()
        self.Show_WeatherInfo2(self.nowinfo_Weather)

    def Show_WeatherInfo1(self) :
        self.canvas_Info.delete('all')

        width = 567
        height = 300

        min = eval(self.allinfo_Weather['0300']['TMP'])
        max = eval(self.allinfo_Weather['0300']['TMP'])

        for i in range(3, 24) :
            temp = eval(self.allinfo_Weather['{t:02}'.format(t= i) + "00"]['TMP'])

            if temp < min :
                min = temp
            
            if temp > max :
                max = temp

        count = len(self.allinfo_Weather)

        x_width = width // (count + 1)
        x_gap = x_width // 5
        x_start = x_width
        y_gap = 50
        y_start = 25
        y_width = height - y_gap * 2 - y_start
        y_stretch = y_width / (max - min)

        for i in range(count) :
            x0 = x_start + i * x_width
            x1 = x0 + x_gap
            y1 = height - y_gap
            y0 = y1 - y_start - (eval(self.allinfo_Weather['{t:02}'.format(t= i + 3) + "00"]['TMP']) - min) * y_stretch

            self.canvas_Info.create_text(x0 + x_gap / 2, y1 + 10, text= '{t:02}'.format(t= i + 3))
            self.canvas_Info.create_rectangle(x0, y0, x1, y1, fill= 'red')
            self.canvas_Info.create_text(x0 + x_gap / 2, y0 - 10, text= self.allinfo_Weather['{t:02}'.format(t= i + 3) + "00"]['TMP'])

    def Show_WeatherInfo2(self, info_pm) :
        self.canvas_Info2.delete('all')

        self.Create_Rectangle_In_Canvas("발표 일자 : " + info_pm["baseDate"], 0, 0, 0)
        self.Create_Rectangle_In_Canvas("발표 시각 : " + info_pm["baseTime"], 0, 1, 0)
        self.Create_Rectangle_In_Canvas("현재 기온 : " + info_pm["T1H"] + " ℃", 1, 0, 0)
        self.Create_Rectangle_In_Canvas("현재 습도 : " + info_pm["REH"] + " %", 1, 1, 0)
        self.Create_Rectangle_In_Canvas("1시간 강수량 : " + info_pm["RN1"] + " mm", 2, 0, 0)
        self.Create_Rectangle_In_Canvas("강수 형태 : " + info_pm["PTY"], 2, 1, 0)
        self.Create_Rectangle_In_Canvas("동서방향 풍속 : " + info_pm["UUU"] + " m/s", 3, 0, 0)
        self.Create_Rectangle_In_Canvas("남북방향 풍속 : " + info_pm["VVV"] + " m/s", 3, 1, 0)
        self.Create_Rectangle_In_Canvas("현재 풍향 : " + info_pm["VEC"] + " °", 4, 0, 0)
        self.Create_Rectangle_In_Canvas("현재 풍속 : " + info_pm["WSD"] + " m/s", 4, 1, 0)

    # Stock Info
    def Search_Stock(self) :
        # Stock
        callbackURL = "http://apis.data.go.kr/1160100/service/GetStockSecuritiesInfoService/getStockPriceInfo"
        params = '?' + urlencode({
            quote_plus("serviceKey"): self.serviceKey,
            quote_plus("numOfRows"): "14",
            quote_plus("pageNo"): "1",
            quote_plus("returnType"): "xml",
            quote_plus("beginBasDt"): cpptime.date_bn(14),
            quote_plus("itmsNm"): self.selected_stock.get()
        })

        url = callbackURL + params
        response_body = urlopen(url).read()
        root = ET.fromstring(response_body.decode('utf-8'))
        items = root.findall(".//item")

        self.allinfo_Stock = []
        for item in items :
            info_Stock = {
                "date": item.findtext("basDt"),
                "name": item.findtext("itmsNm"),
                "category": item.findtext("mrktCtg"),
                "clpr": item.findtext("clpr"),
                "vs": item.findtext("vs"),
                "rate": item.findtext("fltRt"),
                "trade": item.findtext("trqu")
            }

            self.allinfo_Stock.append(info_Stock)

        self.allinfo_Stock = sorted(self.allinfo_Stock, key= lambda k : k['date'], reverse= False)
        self.Show_Stock()

    def Show_Stock(self) :
        self.Show_StockInfo1()
        self.Show_StockInfo2(self.allinfo_Stock[-1])

    def Show_StockInfo1(self) :
        self.canvas_Info.delete('all')

        width = 567
        height = 300

        temp = self.allinfo_Stock
        temp = sorted(temp, key= lambda k : k['clpr'])
        min = eval(temp[0]['clpr'])
        max = eval(temp[-1]['clpr'])
        count = len(self.allinfo_Stock)

        x_width = width // (count + 1)
        x_gap = x_width // 5
        x_start = x_width
        y_gap = 50
        y_start = 25
        y_width = height - y_gap * 2 - y_start
        y_stretch = y_width / (max - min)

        for i in range(count) :
            x0 = x_start + i * x_width
            x1 = x0 + x_gap
            y1 = height - y_gap
            y0 = y1 - y_start - (eval(self.allinfo_Stock[i]['clpr']) - min) * y_stretch

            self.canvas_Info.create_text(x0 + x_gap / 2, y1 + 10, text= self.allinfo_Stock[i]['date'][4:])
            self.canvas_Info.create_rectangle(x0, y0, x1, y1, fill= 'red')
            self.canvas_Info.create_text(x0 + x_gap / 2, y0 - 10, text= self.allinfo_Stock[i]['clpr'])

    def Show_StockInfo2(self, info_pm) :
        self.canvas_Info2.delete('all')

        self.Create_Rectangle_In_Canvas("기준 일자 : " + info_pm["date"], 0, 0, 0)
        self.Create_Rectangle_In_Canvas("종목명 : " + info_pm["name"], 0, 1, 0)
        self.Create_Rectangle_In_Canvas("시장 구분 : " + info_pm["category"], 1, 0, 0)
        self.Create_Rectangle_In_Canvas("종가 : " + info_pm["clpr"] + " ￦", 1, 1, 0)
        self.Create_Rectangle_In_Canvas("전일 대비 : " + info_pm["vs"] + " ￦", 2, 0, self.Return_Color(info_pm["vs"]))
        self.Create_Rectangle_In_Canvas("등락률 : " + info_pm["rate"] + " %", 2, 1, self.Return_Color(info_pm["vs"]))
        self.Create_Rectangle_In_Canvas("거래량 : " + info_pm["trade"], 3, 0, 0)
        
        # grade
        self.Create_Rectangle_In_Canvas("하락", 12, 0, 1)
        self.Create_Rectangle_In_Canvas("상승", 12, 1, 4)

    def Return_Color(self, v) :
        if v[0] == '-' :
            return 1
        else :
            return 4
        
    # telegram
    def replyAptData(self, date_param, user, loc_param='11710'):
        print(user, date_param, loc_param)
        res_list = noti.getData( loc_param, date_param )
        msg = ''
        for r in res_list:
            print( str(datetime.now()).split('.')[0], r )
            if len(r+msg)+1>noti.MAX_MSG_LENGTH:
                noti.sendMessage( user, msg )
                msg = r+'\n'
            else:
                msg += r+'\n'
        if msg:
            noti.sendMessage( user, msg )
        else:
            noti.sendMessage( user, '%s 기간에 해당하는 데이터가 없습니다.'%date_param )

    def save( user, loc_param ):
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS users( user TEXT, location TEXT, PRIMARY KEY(user, location) )')
        try:
            cursor.execute('INSERT INTO users(user, location) VALUES ("%s", "%s")' % (user, loc_param))
        except sqlite3.IntegrityError:
            noti.sendMessage( user, '이미 해당 정보가 저장되어 있습니다.' )
            return
        else:
            noti.sendMessage( user, '저장되었습니다.' )
            conn.commit()

    def check(self, user):
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS users( user TEXT, location TEXT, PRIMARY KEY(user, location) )')
        cursor.execute('SELECT * from users WHERE user="%s"' % user)
        for data in cursor.fetchall():
            row = 'id:' + str(data[0]) + ', location:' + data[1]
            noti.sendMessage( user, row )

    def PM(self, user, name) :
        callbackURL = "http://apis.data.go.kr/B552584/ArpltnInforInqireSvc/getMsrstnAcctoRltmMesureDnsty"
        params = '?' + urlencode({
            quote_plus("serviceKey"): self.serviceKey,
            quote_plus("returnType"): "xml",
            quote_plus("numOfRows"): "10",
            quote_plus("pageNo"): "1",
            quote_plus("stationName"): re.sub('\d+', '', name),
            quote_plus("dataTerm"): "DAILY",
            quote_plus("ver"): "1.4"
        })

        url = callbackURL + params
        response_body = urlopen(url).read()
        root = ET.fromstring(response_body.decode('utf-8'))
        items = root.findall(".//item")            

        temp_PM = []
        for item in items :
            info_PM = {
                "dataTime": item.findtext("dataTime"),
                "stationName": item.findtext("stationName"),
                "pm10": item.findtext("pm10Value"),
                "pm10Grade": item.findtext("pm10Grade"),
                "pm25": item.findtext("pm25Value"),
                "pm25Grade": item.findtext("pm25Grade"),
                "o3": item.findtext("o3Value"),
                "o3Grade": item.findtext("o3Grade"),
                "no2": item.findtext("no2Value"),
                "no2Grade": item.findtext("no2Grade"),
                "co": item.findtext("coValue"),
                "coGrade": item.findtext("coGrade"),
                "so2": item.findtext("so2Value"),
                "so2Grade": item.findtext("so2Grade")
            }

            temp_PM.append(info_PM)

        info = temp_PM[0]

        noti.sendMessage(user, "측정 시간 : " + info["dataTime"])
        noti.sendMessage(user, "측정소 : " + info["stationName"])
        noti.sendMessage(user, "미세먼지 : " + info["pm10"] + " ㎍/㎥")
        noti.sendMessage(user, "초미세먼지 : " + info["pm25"] + " ㎍/㎥")
        noti.sendMessage(user, "오존 : " + info["o3"] + " ppm")
        noti.sendMessage(user, "이산화질소 : " + info["no2"] + " ppm")
        noti.sendMessage(user, "일산화탄소 : " + info["co"] + " ppm")
        noti.sendMessage(user, "아황산가스 : " + info["so2"] + " ppm")

    def Weather(self, user, name) :
        today = cpptime.date_bn(0)
        nowtime = datetime.datetime.now().strftime("%H")

        if nowtime[0] == '0' :
            nowtime = nowtime[0] + str(eval(nowtime[1]) - 1) + "00"
        else :
            nowtime = str(eval(nowtime) - 1) + "00"

        # x, y
        nx, ny = -1, -1

        for row in self.LocalData_ws.rows :
            if row[2].value == name :
                nx, ny = row[3].value, row[4].value
                break

        temp_Weather = dict()

        if nx != -1 and ny != -1 :
            callbackURL = "http://apis.data.go.kr/1360000/VilageFcstInfoService_2.0/getUltraSrtNcst"
            params = '?' + urlencode({
                quote_plus("serviceKey"): self.serviceKey,
                quote_plus("pageNo"): "1",
                quote_plus("numOfRows"): "500",
                quote_plus("dataType"): "XML",
                quote_plus("base_date"): today,
                quote_plus("base_time"): nowtime,
                quote_plus("nx"): nx,
                quote_plus("ny"): ny
            })

            url = callbackURL + params
            response_body = urlopen(url).read()
            root = ET.fromstring(response_body.decode('utf-8'))
            items = root.findall(".//item")

            temp_Weather["baseDate"] = today
            temp_Weather["baseTime"] = nowtime

            for item in items :
                temp_Weather[item.findtext("category")] = item.findtext("obsrValue")

        info = temp_Weather

        noti.sendMessage(user, "발표 일자 : " + info["baseDate"])
        noti.sendMessage(user, "발표 시각 : " + info["baseTime"])
        noti.sendMessage(user, "현재 기온 : " + info["T1H"] + " ℃")
        noti.sendMessage(user, "현재 습도 : " + info["REH"] + " %")
        noti.sendMessage(user, "1시간 강수량 : " + info["RN1"] + " mm")
        noti.sendMessage(user, "강수 형태 : " + info["PTY"])
        noti.sendMessage(user, "동서방향 풍속 : " + info["UUU"] + " m/s")
        noti.sendMessage(user, "남북방향 풍속 : " + info["VVV"] + " m/s")
        noti.sendMessage(user, "현재 풍향 : " + info["VEC"] + " °")
        noti.sendMessage(user, "현재 풍속 : " + info["WSD"] + " m/s")

    def stock(self, user, name) :
        # Stock
        callbackURL = "http://apis.data.go.kr/1160100/service/GetStockSecuritiesInfoService/getStockPriceInfo"
        params = '?' + urlencode({
            quote_plus("serviceKey"): self.serviceKey,
            quote_plus("numOfRows"): "7",
            quote_plus("pageNo"): "1",
            quote_plus("returnType"): "xml",
            quote_plus("beginBasDt"): cpptime.date_bn(7),
            quote_plus("itmsNm"): name
        })

        url = callbackURL + params
        response_body = urlopen(url).read()
        root = ET.fromstring(response_body.decode('utf-8'))
        items = root.findall(".//item")

        temp_Stock = []
        for item in items :
            info_Stock = {
                "date": item.findtext("basDt"),
                "name": item.findtext("itmsNm"),
                "category": item.findtext("mrktCtg"),
                "clpr": item.findtext("clpr"),
                "vs": item.findtext("vs"),
                "rate": item.findtext("fltRt"),
                "trade": item.findtext("trqu")
            }

            temp_Stock.append(info_Stock)
            
        temp_Stock = sorted(temp_Stock, key= lambda k : k['date'], reverse= False)
        info = temp_Stock[-1]
        
        noti.sendMessage(user, "기준 일자 : " + info["date"])
        noti.sendMessage(user, "종목명 : " + info["name"])
        noti.sendMessage(user, "시장 구분 : " + info["category"])
        noti.sendMessage(user, "종가 : " + info["clpr"] + " ￦")
        noti.sendMessage(user, "전일 대비 : " + info["vs"] + " ￦")
        noti.sendMessage(user, "등락률 : " + info["rate"] + " %")
        noti.sendMessage(user, "거래량 : " + info["trade"])

    def handle(self, msg):
        content_type, chat_type, chat_id = telepot.glance(msg)
        if content_type != 'text':
            noti.sendMessage(chat_id, '난 텍스트 이외의 메시지는 처리하지 못해요.')
            return

        text = msg['text']
        args = text.split(' ')
        if text.startswith('미세먼지') and len(args)>1:
            print('try to 미세먼지', args[1])
            self.PM(chat_id, args[1] )
        elif text.startswith('날씨') and len(args)>1:
            print('try to 날씨', args[1])
            self.Weather(chat_id, args[1] )
        elif text.startswith('주식') and len(args)>1:
            print('try to 주식', args[1])
            self.stock(chat_id, args[1] )
        elif text.startswith('확인'):
            print('try to 확인')
            self.check( chat_id )
        else:
            noti.sendMessage(chat_id, '모르는 명령어입니다.\n미세먼지 [동네명]\n날씨 [동네명]\n주식 [종목명]\n확인\n중 하나의 명령을 입력하세요.')


HAMBOGOGA()